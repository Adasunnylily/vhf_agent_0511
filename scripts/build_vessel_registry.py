#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Iterable


GENERIC_SPEAKERS = {
    "船方",
    "船方1",
    "船方2",
    "另一船",
    "交管",
    "宁波交管",
    "待确认",
    "待确认说话人",
    "疑似船方A",
    "疑似船方B",
}


def clean_name(value: object) -> str:
    name = str(value or "").strip()
    name = re.sub(r"^[\[【]|[\]】]$", "", name).strip()
    if not name or name in GENERIC_SPEAKERS or "/" in name:
        return ""
    if len(name) > 32 or not re.search(r"[\u4e00-\u9fffA-Za-z]", name):
        return ""
    return name


def read_workbook_names(path: Path) -> set[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请在构建环境安装后重试") from exc

    names: set[str] = set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value or "").strip() for value in rows[0]]
        name_columns = {
            index
            for index, header in enumerate(headers)
            if header in {"中文船名", "船名", "ship_name", "ship_name_gt"}
        }
        for row in rows[1:]:
            for index in name_columns:
                if index < len(row):
                    name = clean_name(row[index])
                    if name:
                        names.add(name)
            for value in row:
                text = str(value or "")
                for speaker in re.findall(r"[\[【]([^\]】]{2,32})[\]】]", text):
                    name = clean_name(speaker)
                    if name:
                        names.add(name)
                for numbered_name in re.findall(
                    r"(?<![\u4e00-\u9fffA-Za-z0-9])([\u4e00-\u9fffA-Za-z]{2,8}\d{1,5})(?=[，,。\s]|叫|在|向|靠|从|请)",
                    text,
                ):
                    name = clean_name(numbered_name)
                    if name:
                        names.add(name)
    return names


def read_json_names(path: Path, keys: Iterable[str]) -> dict[str, set[str]]:
    aliases: dict[str, set[str]] = {}
    if not path.exists():
        return aliases
    payload = json.loads(path.read_text(encoding="utf-8"))
    for key in keys:
        values = payload.get(key, []) if isinstance(payload, dict) else []
        for item in values:
            if isinstance(item, str):
                canonical = clean_name(item)
                item_aliases: list[str] = []
            elif isinstance(item, dict):
                canonical = clean_name(item.get("canonical"))
                item_aliases = [clean_name(value) for value in item.get("aliases", [])]
            else:
                continue
            if canonical:
                aliases.setdefault(canonical, set()).update(value for value in item_aliases if value)
    return aliases


def build_registry(
    workbooks: list[Path],
    lexicon_path: Path,
    hotwords_path: Path,
    output_path: Path,
) -> dict[str, object]:
    names: set[str] = set()
    for path in workbooks:
        if path.exists():
            names.update(read_workbook_names(path))

    aliases = read_json_names(lexicon_path, ("ships",))
    hotword_aliases = read_json_names(hotwords_path, ("vessels", "ships"))
    for source in (aliases, hotword_aliases):
        names.update(source)

    ships = []
    for canonical in sorted(names, key=lambda value: (value.lower(), value)):
        known_aliases = sorted((aliases.get(canonical, set()) | hotword_aliases.get(canonical, set())) - {canonical})
        ships.append(
            {
                "canonical": canonical,
                "aliases": known_aliases,
                "source": "controlled_registry",
            }
        )

    payload: dict[str, object] = {
        "version": 1,
        "policy": "ship names may only match entries from reviewed workbooks or existing reviewed lexicons",
        "sources": sorted({path.name for path in workbooks if path.exists()}),
        "ships": ships,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the reviewed VHF vessel-name registry.")
    parser.add_argument("--workbook", action="append", default=[], help="Reviewed ship-name or annotation XLSX")
    parser.add_argument("--lexicon", default="data/lexicon_corrections.json")
    parser.add_argument("--hotwords", default="data/hotwords/nbzh_hotwords_llm.json")
    parser.add_argument("--out", default="data/hotwords/nbzh_vessel_registry.json")
    args = parser.parse_args()
    payload = build_registry(
        [Path(value) for value in args.workbook],
        Path(args.lexicon),
        Path(args.hotwords),
        Path(args.out),
    )
    print(f"vessels={len(payload['ships'])} out={Path(args.out).resolve()}")


if __name__ == "__main__":
    main()
